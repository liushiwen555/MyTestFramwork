# -*- coding: utf-8 -*-
# @Time     : 2020/11/25 3:07 下午
# @Author   : LiuShiWen


import traceback
import openpyxl
import os
import pandas as pd
from Common.getLog import logger
from Common.getConfig import Config
from Common.util import stringToDict


class OperateXlsx:
    def __init__(self,filename,sheetname):
        '''
        :type filename: str -- 要打开的文件名
        :param sheetname: str -- 选择操作的excel页名
        '''
        self.logger = logger("error")
        self.filepath = os.path.join(Config().get_option_value('TestDataPath', 'UITestDataPath'), filename)
        if not os.path.exists(self.filepath):
            self.logger.error("文件路径不存在，请检查路径是否正确")
            raise FileNotFoundError
        self.workbook = openpyxl.load_workbook(self.filepath)
        sheetnames = self.workbook.sheetnames
        if sheetname not in sheetnames:
            self.logger.error("请检查传入sheet是否存在")
            raise FileNotFoundError
        self.worksheet = self.workbook[sheetname]

    def maxRow(self):
        """获取excel文件中的总行数"""
        maxrow = self.worksheet.max_row
        return maxrow

    def maxCol(self):
        """获取excel文件中最大的列数"""
        maxcol = self.worksheet.max_column
        return maxcol

    def readByCell(self,row,col):
        """
        获取某个单元格内容
        :param row: 行
        :param col: 列
        :return:
        """
        try:
            info = self.worksheet.cell(row=row, column=col).value
            info = str(info)
            return info
        except:
            self.logger.error("检查传入的行和列格式和数据是否正确")
            self.logger.error(traceback.format_exc())

    def readByRow(self,row):
        """
        获取某一行数据
        :param row: 行号
        :return:
        """
        maxcol=self.maxCol()
        info=[]
        try:
            for i in range(1,maxcol+1):
                content = self.worksheet.cell(row=row,column=i).value
                content = str(content)
                info.append(content)
            return info
        except:
            self.logger.error("检查传入行号的格式和数据是否正确")
            self.logger.error(traceback.format_exc())

    def readByCol(self,column):
        """
        获取某一列数据
        :param column: 列号
        :return:
        """
        maxrow = self.maxRow()
        info=[]
        try:
            for i in range(1, maxrow + 1):
                content = self.worksheet.cell(row=i, column=column).value
                content = str(content)
                info.append(content)
            return info
        except:
            self.logger.error("请检查传入列号的格式和具体数据是否正确")
            self.logger.error(traceback.format_exc())

    @property
    def readByRowAll(self):
        """按行获取所有数据"""
        maxrow=self.maxRow()
        info=[]
        for i in range(1,maxrow+1):
            content = self.readByRow(i)
            info.append(content)
        return info

    def writeByCell(self, rownum, colnum, data):
        """
        在指定单元格写数据
        :param rownum: 行号
        :param colnum: 列号
        :param data: 数据
        :return:
        """
        try:
            self.worksheet.cell(row=rownum, column=colnum).value = data
            self.workbook.save(self.filepath)
        except:
            self.logger.error("检查传入数据是否正确")
            self.logger.error(traceback.format_exc())

    def writeByRow(self, row, data):
        """
        在指定行写一行内容，data为列表
        :param row: 行号
        :param data: 数据，type is list
        :return:
        """
        try:
            num=len(data)
            for i in range(0,num):
                self.worksheet.cell(row=row, column=i+1).value=data[i]
            self.workbook.save(self.filepath)
        except:
            self.logger.error("检查传入数据是否正确")
            self.logger.error(traceback.format_exc())


if __name__ == '__main__':
    work = OperateXlsx('test_data.xlsx', 'Register')
    test_data = work.readByRow(2)[4]
    test_data_by_col = work.readByCol(5)[1]
    test_data = stringToDict(test_data)
    test_data_by_col = stringToDict(test_data_by_col)
    print(test_data,"\n",type(test_data))
    print(test_data_by_col)












# class ReadEmailXlsx(object):
#     def __init__(self,filepath):
#         self.book = openpyxl.load_workbook(filepath)
#         self.sheet = self.book.active
#         self.rows_num = self.sheet.max_row
#         self.cols_num = self.sheet.max_column
#     def read_by_colums(self)->list:
#         read_by_cols = []
#         for col in self.sheet.iter_cols(max_col=self.cols_num):
#             for cell in col:
#                 read_cell = cell.value
#                 read_by_cols.append(read_cell)
#                 # print(read_by_cols)
#         return read_by_cols
#
# if __name__ == '__main__':
#     filepath = os.path.join(Config().get_option_value('TestDataPath','UITestDataPath'),'test_data.xlsx')
#     # print(filepath)
#     print(ReadEmailXlsx(filepath).read_by_colums())

# class OperateXlsx(object):
#
#     def __init__(self,filename,sheetname):
#         self.filepath = os.path.join(Config().get_option_value('TestDataPath', 'UITestDataPath'), filename)
#         self.workbook = openpyxl.load_workbook(self.filepath)
#         self.worksheet= self.workbook
#
#
#     def readByCells(self):
#         pass
#
#     def readByRows(self):
#         pass
#
#     def readByCols(self):
#         pass
#
#
# if __name__ == '__main__':
#     filepath = os.path.join(Config().get_option_value('TestDataPath', 'UITestDataPath'), 'test_data.xlsx')
#     workbook = openpyxl.load_workbook(filepath)
#     sheetnames = workbook.sheetnames
#     print(sheetnames)
#     worksheet = workbook[sheetnames[0]]
#     print(worksheet.max_row)
#     print(worksheet.max_column)